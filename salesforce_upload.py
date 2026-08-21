from __future__ import annotations

from datetime import datetime, timedelta
from typing import Dict, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SALESFORCE_COLUMNS = [
    "Work Order Number",
    "Primary Service Appointment",
    "Service Appointment ID",
    "Service Resource ID",
    "Resource",
    "Scheduled Start",
    "Scheduled End",
    "Duration (Minutes)",
    "Customer Reference Code",
    "Building Name",
]

# Only these surveyors are currently approved for the Salesforce upload tab.
# Harrison Grice and Joe Reynolds intentionally remain excluded until their
# Service Resource IDs are known.
SALESFORCE_RESOURCES: Dict[str, Tuple[str, str]] = {
    "Conor Birch": ("Harpenden", "0HnR50000005RlxKAE"),
    "Rod Harrison": ("Rugby", "0Hn4L0000000Yy8SAE"),
    "Toby Lawal": ("Chadwell Heath", "0HnR50000005S6vKAE"),
}


def _clean_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.endswith(".0"):
        integer_part = text[:-2]
        if integer_part.lstrip("-").isdigit():
            return integer_part
    return text


def _work_order_text(value) -> str:
    """Preserve Salesforce's leading-zero Work Order format."""
    text = _clean_text(value)
    if not text:
        return ""
    if text.isdigit() and len(text) < 8:
        return text.zfill(8)
    return text


def _scheduled_datetime(date_value, time_value):
    date_text = _clean_text(date_value)
    time_text = _clean_text(time_value)
    if not date_text or not time_text:
        return None

    parsed = pd.to_datetime(
        f"{date_text} {time_text}",
        errors="coerce",
    )
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def _salesforce_timestamp(dt) -> str:
    if dt is None:
        return ""
    # Match the user's supplied Salesforce copy template exactly:
    # 2026-08-12T09:00:00.000+0000
    return dt.strftime("%Y-%m-%dT%H:%M:%S.000+0000")


def build_salesforce_upload_dataframe(
    combined_team_schedule: pd.DataFrame,
    team_portfolio: pd.DataFrame,
) -> pd.DataFrame:
    """
    Create the copy/paste Salesforce Field Service upload rows.

    Source of truth:
      - schedule: Surveyor, Date, Survey Start, Survey End
      - To Do/master portfolio: Work Order Number, Primary Service Appointment,
        Service Appointment ID, Customer Reference, Building Name

    Only Conor Birch, Rod Harrison and Toby Lawal are included.
    """
    if combined_team_schedule is None or combined_team_schedule.empty:
        return pd.DataFrame(columns=SALESFORCE_COLUMNS)

    if team_portfolio is None or team_portfolio.empty:
        return pd.DataFrame(columns=SALESFORCE_COLUMNS)

    portfolio = team_portfolio.copy()
    if "Customer Reference" not in portfolio.columns:
        return pd.DataFrame(columns=SALESFORCE_COLUMNS)

    portfolio["__sf_ref"] = portfolio["Customer Reference"].map(_clean_text)
    portfolio = portfolio[portfolio["__sf_ref"].ne("")].copy()
    portfolio = portfolio.drop_duplicates(subset=["__sf_ref"], keep="first")
    lookup = portfolio.set_index("__sf_ref", drop=False)

    rows = []
    for _, scheduled in combined_team_schedule.iterrows():
        surveyor = _clean_text(scheduled.get("Surveyor"))
        if surveyor not in SALESFORCE_RESOURCES:
            continue

        customer_ref = _clean_text(scheduled.get("Customer Reference"))
        if not customer_ref or customer_ref.upper() == "RETURN":
            continue
        if customer_ref not in lookup.index:
            # Never invent Salesforce identifiers. A row without a matching
            # To Do record is omitted from the copy/paste upload sheet.
            continue

        source = lookup.loc[customer_ref]
        if isinstance(source, pd.DataFrame):
            source = source.iloc[0]

        resource_name, resource_id = SALESFORCE_RESOURCES[surveyor]

        start_dt = _scheduled_datetime(
            scheduled.get("Date"),
            scheduled.get("Survey Start"),
        )
        end_dt = _scheduled_datetime(
            scheduled.get("Date"),
            scheduled.get("Survey End"),
        )
        if start_dt is not None and end_dt is not None and end_dt < start_dt:
            end_dt = end_dt + timedelta(days=1)

        if start_dt is not None and end_dt is not None:
            duration_minutes = int(
                round((end_dt - start_dt).total_seconds() / 60.0)
            )
        else:
            try:
                duration_minutes = int(
                    round(float(scheduled.get(
                        "Planning Survey Duration (Minutes)", 0
                    )))
                )
            except (TypeError, ValueError):
                duration_minutes = ""

        rows.append({
            "Work Order Number": _work_order_text(
                source.get("Work Order Number")
            ),
            "Primary Service Appointment": _clean_text(
                source.get("Primary Service Appointment: Appointment Number")
            ),
            "Service Appointment ID": _clean_text(
                source.get(
                    "Primary Service Appointment: Service Appointment ID"
                )
            ),
            "Service Resource ID": resource_id,
            "Resource": resource_name,
            "Scheduled Start": _salesforce_timestamp(start_dt),
            "Scheduled End": _salesforce_timestamp(end_dt),
            "Duration (Minutes)": duration_minutes,
            "Customer Reference Code": customer_ref,
            "Building Name": _clean_text(source.get("Building Name")),
        })

    return pd.DataFrame(rows, columns=SALESFORCE_COLUMNS)


def write_salesforce_upload_sheet(
    workbook,
    upload_df: pd.DataFrame,
    sheet_name: str = "Salesforce Upload",
    index: int = 2,
):
    """Write and style a worksheet to match the supplied Salesforce template."""
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    ws = workbook.create_sheet(title=sheet_name, index=index)

    dark_blue = "17365D"
    white = "FFFFFF"
    text_colour = "1F2937"
    yellow = "FFFF00"
    header_border_colour = "AAB7C4"
    row_border_colour = "E5E7EB"

    blue_fill = PatternFill(fill_type="solid", fgColor=dark_blue)
    yellow_fill = PatternFill(fill_type="solid", fgColor=yellow)

    header_side = Side(style="thin", color=header_border_colour)
    row_side = Side(style="thin", color=row_border_colour)

    # Row 1: exact title area from the supplied template.
    for col in range(1, 11):
        cell = ws.cell(row=1, column=col)
        cell.fill = blue_fill
    ws.merge_cells("A1:J1")
    title = ws["A1"]
    title.value = "Salesforce Field Service Upload"
    title.font = Font(name="Carlito", size=16, bold=True, color=white)
    ws.row_dimensions[1].height = 20.25

    # Row 2 intentionally blank.

    # Row 3: field names in the exact supplied order.
    for col_idx, heading in enumerate(SALESFORCE_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=heading)
        cell.fill = blue_fill
        cell.font = Font(name="Carlito", size=11, bold=True, color=white)
        cell.alignment = Alignment(wrap_text=True)
        left = header_side if col_idx == 1 else Side(style=None)
        right = header_side if col_idx == 10 else Side(style=None)
        cell.border = Border(
            top=header_side,
            bottom=header_side,
            left=left,
            right=right,
        )
    ws.row_dimensions[3].height = 30

    # Detail rows.
    if upload_df is None:
        upload_df = pd.DataFrame(columns=SALESFORCE_COLUMNS)

    for row_offset, record in enumerate(
        upload_df.reindex(columns=SALESFORCE_COLUMNS).to_dict(orient="records"),
        start=4,
    ):
        for col_idx, heading in enumerate(SALESFORCE_COLUMNS, start=1):
            value = record.get(heading, "")
            if pd.isna(value):
                value = ""
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.font = Font(name="Carlito", size=11, color=text_colour)
            if col_idx in (3, 4):
                cell.fill = yellow_fill
            cell.border = Border(
                top=(row_side if row_offset > 4 else Side(style=None)),
                bottom=row_side,
            )


    # The template uses the same width across all 10 fields.
    for col_idx in range(1, 11):
        ws.column_dimensions[get_column_letter(col_idx)].width = 26.140625

    return ws
