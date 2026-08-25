import math
from pathlib import Path
from datetime import datetime, time, timedelta
from zoneinfo import ZoneInfo
import io
import os

import pandas as pd
import streamlit as st

from duration_predictor_height import DurationPredictor, FEATURE_COLUMNS
from google_routes import GoogleTransitRouter, GoogleRoutesError
from scheduler_v20_10 import DailyTransitScheduler, postcode_district
from coordinate_clustering import (
    GEOGRAPHIC_CLUSTER_MAX_DIAMETER_KM,
    NO_GOOGLE_RADIUS_KM,
)
import inspect

_REQUIRED_BUILD_WEEK_ARGS = {
    "first_survey_start_clock",
    "latest_survey_finish_clock",
    "latest_return_clock",
}
_build_week_args = set(
    inspect.signature(DailyTransitScheduler.build_week).parameters
)
if not _REQUIRED_BUILD_WEEK_ARGS.issubset(_build_week_args):
    raise RuntimeError(
        "Version 20.10 scheduler mismatch: replace scheduler_v20_10.py and "
        "the other Version 20.10 changed files, then reboot the app."
    )

from ai_planner import OpenAISchedulePlanner
from tfl_client import TfLClient
from metoffice_client import MetOfficeClient
from salesforce_master import read_salesforce_or_standard_excel
from special_requests import (
    SpecialRequestResult,
    all_cluster_representatives,
    choose_nearby_cluster,
    scheduled_reference_set,
    build_trial_dataframe,
    request_results_dataframe,
)
from team_scheduler import (
    SurveyorConfig,
    representative_sites,
    home_to_cluster_matrix,
    allocate_cluster_targets,
    build_team_shortlists,
    allocations_dataframe,
)
from portfolio_clusterer import (
    add_portfolio_fields,
    build_cluster_summary,
    deterministic_cluster_choices,
    shortlist_sites,
    build_drawing_priority_queue,
    endgame_adjust_cluster_choices,
)


LONDON_TZ = ZoneInfo("Europe/London")
DEFAULT_FILE = Path(__file__).with_name("Predictive Model.xlsx")

st.set_page_config(page_title="Site Survey Scheduling Agent", layout="wide")
st.title("Site Survey Scheduling Agent")
st.caption("Version 20.10.2 — Salesforce upload includes Harrison + Joe")
st.caption(
    "Upload the master portfolio, set surveyor availability for one week, then "
    "use Google transit routing only for that selected week."
)

with st.sidebar:
    st.header("Duration model")
    training_file = st.file_uploader(
        "Completed-surveys training spreadsheet",
        type=["xlsx", "xls"],
        key="training_file",
        help="Supports both normal tables and Salesforce report exports with title/filter rows above the headers. If omitted, the bundled Predictive Model.xlsx is used.",
    )
    min_duration = st.number_input(
        "Minimum completed duration (minutes)",
        min_value=1,
        max_value=30,
        value=6,
        help="Historical rows below this are excluded from model training.",
    )
    st.caption(
        "Planning duration now equals the raw model prediction; no survey-duration "
        "percentage uplift is applied."
    )

@st.cache_resource(show_spinner=False)
def train_from_path_v20_2(path: str, min_duration: int):
    return DurationPredictor(
        min_completed_duration=min_duration,
    ).load_excel(path)

def read_completed_training_excel(source, sheet_name=0):
    """
    Read either the original row-1 completed-surveys table or a Salesforce
    report export with title/filter rows above the real headers.

    This parser lives in app.py deliberately so the upload path does not depend
    on a particular DurationPredictor class version being cached/deployed.
    """
    raw = pd.read_excel(source, sheet_name=sheet_name, header=None)

    target_aliases = {
        "Primary Service Appointment: Actual Duration (Minutes)",
        "Primary Service Appointment: Actual Duration",
        "Actual Duration (Minutes)",
    }
    header_aliases = {
        "Building Height",
        "Building height",
        "Internal Ground Floor Area (m2)",
        "Internal Ground Floor Area",
        "Internal Ground Floor Area (m²)",
        "Sovereign Flat",
        "Sovereign Flats",
        "Sovereign Flat Count",
        *target_aliases,
    }

    header_row = None
    for idx, row in raw.iterrows():
        values = {str(v).strip() for v in row.tolist() if pd.notna(v)}
        if values.intersection(target_aliases) and len(values.intersection(header_aliases)) >= 3:
            header_row = int(idx)
            break

    if header_row is None or header_row == 0:
        return pd.read_excel(source, sheet_name=sheet_name)

    df = pd.read_excel(source, sheet_name=sheet_name, header=header_row)
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")

    # Remove Salesforce Total / Sum / Count report footer rows.
    text_view = df.astype(str).apply(lambda col: col.str.strip().str.lower())
    footer_mask = text_view.apply(
        lambda row: row.isin({"total", "sum", "count"}).any(), axis=1
    )
    df = df[~footer_mask].copy()

    # Keep actual detail rows, not any remaining report summaries.
    identity_names = {
        "Building Name",
        "Customer Reference Code  ↑",
        "Customer Reference Code",
        "Customer Reference",
        "Work Order Number",
    }
    identity_cols = [c for c in df.columns if str(c).strip() in identity_names]
    if identity_cols:
        identity_present = pd.Series(False, index=df.index)
        for col in identity_cols:
            values = df[col]
            identity_present |= values.notna() & values.astype(str).str.strip().ne("")
        df = df[identity_present].copy()

    return df.reset_index(drop=True)


@st.cache_resource(show_spinner=False)
def train_from_bytes_v20_2(file_bytes: bytes, min_duration: int):
    df = read_completed_training_excel(io.BytesIO(file_bytes))
    return DurationPredictor(
        min_completed_duration=min_duration,
    ).fit(df)

try:
    if training_file is not None:
        predictor = train_from_bytes_v20_2(
            training_file.getvalue(), min_duration
        )
    else:
        predictor = train_from_path_v20_2(
            str(DEFAULT_FILE), min_duration
        )
except Exception as exc:
    st.error(f"Could not train duration model: {exc}")
    st.stop()

# Fail loudly rather than silently reusing an older cached/unsegmented predictor.
if getattr(predictor, "model_version", None) != "20.2-segmented-1-6":
    st.error(
        "Old duration predictor detected. Replace all repository files with "
        "Version 20.2 and reboot the Streamlit app."
    )
    st.stop()

with st.sidebar:
    st.success("Active model: v20.2 segmented (Garage / 1–6 flats / 7+ flats)")


def optional_number(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if pd.isna(result):
        return None
    return result


def normalise_upcoming_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = predictor._normalise_columns(df.copy())

    # Exact Salesforce export aliases used in Future Surveys.xlsx.
    aliases = {
        "Customer Reference": [
            "Customer Reference Code  ↑",
            "Customer Reference Code",
            "Customer Reference",
        ],
        "Building Name": ["Building Name"],
        "Postcode": ["Postcode", "Postal Code"],
        "Resource Name": ["Resource Name: Name", "Resource Name"],
        "Planned Start": [
            "Planned Start",
            "Primary Service Appointment: Scheduled Start",
        ],
        "Drawing Status": [
            "Drawing Status",
            "Draw Status",
            "Drawing State",
        ],
        "Earliest Survey Date": [
            "Earliest Survey Date",
            "Earliest Survey",
            "Available From",
        ],
    }

    rename = {}
    stripped = {str(c).strip(): c for c in df.columns}
    for canonical, possibilities in aliases.items():
        for option in possibilities:
            if option in stripped:
                rename[stripped[option]] = canonical
                break

    return df.rename(columns=rename)


def predict_upcoming(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in df.iterrows():
        try:
            p = predictor.predict(
                building_height=row.get(
                    FEATURE_COLUMNS["building_height"]
                ),
                ground_floor_area=row.get(
                    FEATURE_COLUMNS["ground_floor_area"]
                ),
                flats=row.get(FEATURE_COLUMNS["flats"]),
            )
            rows.append({
                "Predicted Survey Duration (Minutes)": p.predicted_minutes,
                "Planning Duration (Minutes)": p.planning_minutes,
                "Prediction Confidence": p.confidence,
                "Prediction Model Used": p.model_label,
                "Validation MAE (Minutes)": p.validation_mae_minutes,
                "Validation RMSE (Minutes)": p.validation_rmse_minutes,
                "Prediction Training Rows": p.training_rows,
                "Prediction Feature Count": p.feature_count,
                "Prediction Status": "Predicted",
            })
        except ValueError:
            rows.append({
                "Predicted Survey Duration (Minutes)": None,
                "Planning Duration (Minutes)": None,
                "Prediction Confidence": "None",
                "Prediction Model Used": "No usable input data",
                "Validation MAE (Minutes)": None,
                "Validation RMSE (Minutes)": None,
                "Prediction Training Rows": None,
                "Prediction Feature Count": None,
                "Prediction Status": "Could not predict",
            })

    return pd.concat(
        [df.reset_index(drop=True), pd.DataFrame(rows)],
        axis=1,
    )



def get_secret(name: str, default: str = "") -> str:
    try:
        return str(st.secrets[name])
    except (KeyError, FileNotFoundError):
        return os.getenv(name, default)


def planned_date_from_value(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def build_future_cluster_context(df: pd.DataFrame, week_start):
    """
    Count how many candidate sites share a postcode district in this week,
    next week, and the following three weeks. Existing Planned Start is used
    as a planning signal, not as a hard appointment constraint.
    """
    records = []
    for _, row in df.iterrows():
        pc = str(row.get("Postcode", "")).strip()
        d = postcode_district(pc)
        pd_date = planned_date_from_value(row.get("Planned Start"))
        records.append((d, pd_date))

    def count_for(district, start, end):
        return sum(
            1 for d, dt in records
            if d == district and dt is not None and start <= dt <= end
        )

    this_end = week_start + timedelta(days=6)
    next_start = week_start + timedelta(days=7)
    next_end = week_start + timedelta(days=13)
    three_end = week_start + timedelta(days=27)

    rows = []
    summary = {}
    districts = sorted({d for d, _ in records if d})
    for d in districts:
        this_count = count_for(d, week_start, this_end)
        next_count = count_for(d, next_start, next_end)
        next3_count = count_for(d, next_start, three_end)
        summary[d] = {
            "this_week": this_count,
            "next_week": next_count,
            "next_3_weeks": next3_count,
        }
        rows.append(
            f"{d}: this week={this_count}, next week={next_count}, "
            f"next 3 weeks={next3_count}"
        )
    return summary, "\n".join(rows)


def apply_ai_decisions(sites, decisions):
    by_ref = {d.customer_reference: d for d in decisions}
    for site in sites:
        ref = str(site.get("customer_reference", ""))
        decision = by_ref.get(ref)
        if decision:
            site["ai_priority"] = decision.priority
            site["ai_decision"] = decision.decision
            site["ai_reason"] = decision.reason
        else:
            site["ai_priority"] = 50
            site["ai_decision"] = "neutral"
            site["ai_reason"] = "No specific AI decision returned."
    return sites

def site_dataframe_to_dicts(df: pd.DataFrame):
    sites = []
    for _, site_row in df.iterrows():
        building_name = str(site_row.get("Building Name", "")).strip()
        postcode = str(site_row.get("Postcode", "")).strip()
        route_location = (
            f"{building_name}, {postcode}" if building_name else postcode
        )

        special_date = site_row.get("Special Request Date")
        if special_date is not None and not pd.isna(special_date):
            parsed_special_date = pd.to_datetime(
                special_date, errors="coerce"
            )
            special_date = (
                parsed_special_date.date()
                if not pd.isna(parsed_special_date)
                else None
            )
        else:
            special_date = None

        sites.append({
            "customer_reference": site_row.get("Customer Reference", ""),
            "building_name": building_name or postcode,
            "postcode": postcode,
            "postcode_district": postcode_district(postcode),
            "planning_cluster": str(
                site_row.get("Planning Cluster", "")
                or postcode_district(postcode)
            ),
            "latitude": optional_number(site_row.get("Latitude Clean", site_row.get("Latitude"))),
            "longitude": optional_number(site_row.get("Longitude Clean", site_row.get("Longitude"))),
            "route_location": route_location,
            "planning_minutes": float(
                site_row["Planning Duration (Minutes)"]
            ),
            "predicted_minutes": float(
                site_row["Predicted Survey Duration (Minutes)"]
            ),
            "confidence": site_row["Prediction Confidence"],
            "model_used": site_row["Prediction Model Used"],
            "building_height": optional_number(
                site_row.get(FEATURE_COLUMNS["building_height"])
            ),
            "flats": optional_number(
                site_row.get(FEATURE_COLUMNS["flats"])
            ),
            "ground_floor_area": optional_number(
                site_row.get(FEATURE_COLUMNS["ground_floor_area"])
            ),
            "planned_start": str(
                site_row.get("Planned Start", "") or ""
            ),
            "ai_priority": float(
                site_row.get("AI Cluster Priority", 50)
            ),
            "ai_decision": "neutral",
            "ai_reason": str(
                site_row.get("AI Cluster Reason", "")
            ),
            "special_request_date": special_date,
            "special_request_text": str(
                site_row.get("Special Request Text", "") or ""
            ),
            "special_request_target_cluster": str(
                site_row.get("Special Request Target Cluster", "") or ""
            ),
            "special_request_bonus_minutes": 75.0,
        })
    return sites


def result_uses_cluster_on_date(result, requested_date, cluster):
    if result is None:
        return False
    for day in result.days:
        if day.start_time.date() != requested_date:
            continue
        if any(str(item.cluster) == str(cluster) for item in day.items):
            return True
    return False


SALESFORCE_RESOURCE_MAP = {
    "conor birch": {
        "service_resource_id": "0HnR50000005RlxKAE",
        "resource": "Harpenden",
    },
    "rod harrison": {
        "service_resource_id": "0Hn4L0000000Yy8SAE",
        "resource": "Rugby",
    },
    "toby lawal": {
        "service_resource_id": "0HnR50000005S6vKAE",
        "resource": "Chadwell Heath",
    },
    "harrison grice": {
        "service_resource_id": "",
        "resource": "",
    },
    "joe reynolds": {
        "service_resource_id": "",
        "resource": "",
    },
}


def _first_portfolio_value(row, candidates):
    for column in candidates:
        if column not in row.index:
            continue
        value = row.get(column)
        if value is None:
            continue
        try:
            if pd.isna(value):
                continue
        except Exception:
            pass
        text_value = str(value).strip()
        if text_value and text_value.lower() not in {"nan", "none"}:
            return text_value
    return ""


def _salesforce_iso_datetime(date_value, time_value):
    parsed_date = pd.to_datetime(date_value, errors="coerce")
    time_text = str(time_value or "").strip()

    if pd.isna(parsed_date) or not time_text:
        return ""

    combined = pd.to_datetime(
        f"{parsed_date.date().isoformat()} {time_text}",
        errors="coerce",
    )
    if pd.isna(combined):
        return ""

    # Salesforce upload template format supplied by the user.
    return combined.strftime("%Y-%m-%dT%H:%M:00.000+0000")


def build_salesforce_copy(
    schedule_df: pd.DataFrame,
    portfolio_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Build the exact 10-column Salesforce Field Service upload table.

    Conor Birch, Rod Harrison, Toby Lawal, Harrison Grice and Joe Reynolds
    are included. Harrison Grice and Joe Reynolds currently export with blank
    Service Resource ID and Resource fields until those Salesforce values are
    known. LUNCH / RETURN operational rows are excluded automatically.

    Work-order / appointment identifiers and building fields are joined back
    to the uploaded To Do portfolio using Customer Reference Code.
    """
    columns = [
        "Work Order Number",
        "Primary Service Appointment",
        "Service Appointment ID",
        "Service Resource ID",
        "Resource",
        "Scheduled Start",
        "Scheduled End",
        "Duration (Minutes)",
        "Customer Reference Code",
        "Building Name",
    ]

    if schedule_df is None or schedule_df.empty:
        return pd.DataFrame(columns=columns)

    working = schedule_df.copy()

    # Real surveys have a numeric sequence. LUNCH / RETURN do not.
    numeric_sequence = pd.to_numeric(
        working.get("Sequence"),
        errors="coerce",
    )
    working = working[numeric_sequence.notna()].copy()

    if working.empty:
        return pd.DataFrame(columns=columns)

    # Only surveyors explicitly configured for the Salesforce upload are included.
    # Harrison Grice and Joe Reynolds are intentionally configured with blank
    # Resource / Service Resource ID values until those Salesforce details are known.
    working["_surveyor_key"] = (
        working["Surveyor"].fillna("").astype(str).str.strip().str.lower()
    )
    working = working[
        working["_surveyor_key"].isin(SALESFORCE_RESOURCE_MAP)
    ].copy()

    if working.empty:
        return pd.DataFrame(columns=columns)

    # Build a lookup from the uploaded To Do portfolio. normalise_upcoming_columns()
    # turns Customer Reference Code into Customer Reference but preserves all the
    # other original Salesforce columns.
    portfolio_lookup = {}
    if portfolio_df is not None and not portfolio_df.empty:
        ref_col = None
        for candidate in ["Customer Reference", "Customer Reference Code"]:
            if candidate in portfolio_df.columns:
                ref_col = candidate
                break

        if ref_col is not None:
            for _, prow in portfolio_df.iterrows():
                ref = str(prow.get(ref_col, "") or "").strip()
                if ref and ref.lower() not in {"nan", "none"}:
                    portfolio_lookup[ref] = prow

    rows = []
    for _, row in working.iterrows():
        ref = str(row.get("Customer Reference", "") or "").strip()
        portfolio_row = portfolio_lookup.get(ref)

        if portfolio_row is None:
            # Keep the exact structure even if a source row cannot be rejoined.
            portfolio_row = pd.Series(dtype=object)

        resource_info = SALESFORCE_RESOURCE_MAP[row["_surveyor_key"]]

        work_order = _first_portfolio_value(
            portfolio_row,
            [
                "Work Order Number",
                "Work Order",
            ],
        )

        primary_service_appointment = _first_portfolio_value(
            portfolio_row,
            [
                "Primary Service Appointment: Appointment Number",
                "Primary Service Appointment",
                "Appointment Number",
            ],
        )

        service_appointment_id = _first_portfolio_value(
            portfolio_row,
            [
                "Service Appointment ID",
                "Primary Service Appointment: Service Appointment ID",
                "Primary Service Appointment ID",
                "Primary Service Appointment: ID",
                "Service Appointment: ID",
            ],
        )

        building_name = _first_portfolio_value(
            portfolio_row,
            ["Building Name"],
        ) or str(row.get("Building Name", "") or "").strip()

        scheduled_start = _salesforce_iso_datetime(
            row.get("Date"),
            row.get("Survey Start"),
        )
        scheduled_end = _salesforce_iso_datetime(
            row.get("Date"),
            row.get("Survey End"),
        )

        # The Salesforce template uses whole-minute duration. Derive this from
        # the scheduled start/end clock values so all three fields agree exactly.
        duration_minutes = ""
        start_dt = pd.to_datetime(
            f"{row.get('Date')} {row.get('Survey Start')}",
            errors="coerce",
        )
        end_dt = pd.to_datetime(
            f"{row.get('Date')} {row.get('Survey End')}",
            errors="coerce",
        )
        if not pd.isna(start_dt) and not pd.isna(end_dt):
            duration_minutes = int(
                round((end_dt - start_dt).total_seconds() / 60.0)
            )

        rows.append({
            "Work Order Number": work_order,
            "Primary Service Appointment": primary_service_appointment,
            "Service Appointment ID": service_appointment_id,
            "Service Resource ID": resource_info["service_resource_id"],
            "Resource": resource_info["resource"],
            "Scheduled Start": scheduled_start,
            "Scheduled End": scheduled_end,
            "Duration (Minutes)": duration_minutes,
            "Customer Reference Code": ref,
            "Building Name": building_name,
        })

    return pd.DataFrame(rows, columns=columns)


def write_salesforce_copy_sheet(writer, salesforce_copy_df: pd.DataFrame):
    """
    Write the Salesforce Copy worksheet in the exact visual/table layout of
    Salesforce Copy Format Chat(2).xlsx.
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    workbook = writer.book

    if "Salesforce Copy" in workbook.sheetnames:
        del workbook["Salesforce Copy"]

    ws = workbook.create_sheet("Salesforce Copy")

    dark_blue = "17365D"
    id_header_blue = "0E2841"
    yellow = "FFFF00"
    body_text = "1F2937"
    header_border_colour = "AAB7C4"
    body_border_colour = "E5E7EB"

    # Exact template dimensions/layout.
    for col_idx in range(1, 11):
        ws.column_dimensions[get_column_letter(col_idx)].width = 26.140625

    ws.row_dimensions[1].height = 20.25
    ws.row_dimensions[3].height = 30

    ws.merge_cells("A1:J1")
    title = ws["A1"]
    title.value = "Salesforce Field Service Upload"
    title.fill = PatternFill("solid", fgColor=dark_blue)
    title.font = Font(
        name="Carlito",
        size=16,
        bold=True,
        color="FFFFFF",
    )
    title.alignment = Alignment(vertical="center")

    # The source template highlights C/D because these are Salesforce IDs.
    for cell_ref in ("C2", "D2"):
        ws[cell_ref].fill = PatternFill("solid", fgColor=yellow)

    headers = [
        "Work Order Number",
        "Primary Service Appointment",
        "Service Appointment ID",
        "Service Resource ID",
        "Resource",
        "Scheduled Start",
        "Scheduled End",
        "Duration (Minutes)",
        "Customer Reference Code",
        "Building Name",
    ]

    thin_header = Side(style="thin", color=header_border_colour)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(
            name="Carlito",
            size=11,
            bold=True,
            color="FFFFFF",
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=(id_header_blue if col_idx in (3, 4) else dark_blue),
        )
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(
            left=(thin_header if col_idx == 1 else Side()),
            right=(thin_header if col_idx == 10 else Side()),
            top=thin_header,
            bottom=thin_header,
        )

    thin_body = Side(style="thin", color=body_border_colour)

    for row_offset, values in enumerate(
        salesforce_copy_df.itertuples(index=False, name=None),
        start=4,
    ):
        first_data_row = row_offset == 4

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.font = Font(
                name="Carlito",
                size=11,
                color=body_text,
            )
            cell.alignment = Alignment(vertical="center")

            # IDs in C/D are highlighted exactly like the supplied template.
            if col_idx in (3, 4):
                cell.fill = PatternFill("solid", fgColor=yellow)

            # Salesforce treats everything except Duration as literal text.
            if col_idx != 8:
                cell.number_format = "@"

            cell.border = Border(
                top=(Side() if first_data_row else thin_body),
                bottom=thin_body,
            )

    return ws


tab1, tab2, tab3 = st.tabs([
    "Predict one building",
    "Weekly scheduling",
    "Model diagnostics",
])

with tab1:
    st.subheader("Single-building prediction")
    st.write("Leave any unavailable input blank.")

    c1, c2, c3 = st.columns(3)
    with c1:
        height_text = st.text_input(
            "Building Height", placeholder="e.g. 24"
        )
    with c2:
        area_text = st.text_input(
            "Internal Ground Floor Area (m²)", placeholder="e.g. 520"
        )
    with c3:
        flats_text = st.text_input(
            "Sovereign Flats", placeholder="e.g. 42"
        )

    if st.button("Predict survey duration", type="primary"):
        try:
            p = predictor.predict(
                building_height=optional_number(height_text),
                ground_floor_area=optional_number(area_text),
                flats=optional_number(flats_text),
            )
            a, b, c = st.columns(3)
            a.metric("Predicted duration", f"{p.predicted_minutes:.0f} min")
            b.metric("Planning duration", f"{p.planning_minutes} min")
            c.metric("Confidence", p.confidence)
            st.success(f"Model used: {p.model_label}")
        except ValueError as exc:
            st.warning(str(exc))


with tab2:
    st.subheader("Weekly scheduling")
    st.write(
        "Create the selected week's schedules for one or more surveyors. "
        "Every postcode-bearing row in the uploaded portfolio is considered for "
        "forward-looking clustering/endgame planning, while Google Routes is only "
        "used for the one selected week."
    )
    st.caption(
        "Weekly schedule rule: every Work Type and Status in the uploaded portfolio "
        "is considered for the actual schedule — including Plan Drafting, Work Request, "
        "Work Done, Under Preparation and Released. A usable postcode, drawing/date "
        "eligibility and a duration prediction are still required before Google routing."
    )

    team_file = st.file_uploader(
        "Master portfolio spreadsheet",
        type=["xlsx", "xls"],
        key="team_master_file",
        help=(
            "Can contain both survey-ready and Needs Drawing sites. "
            "Recommended columns include Customer Reference, Building Name, "
            "Postcode, Building Height, Sovereign Flat, Internal Ground Floor "
            "Area, Drawing Status and Earliest Survey Date."
        ),
    )

    if team_file is None:
        st.info(
            "Upload the master portfolio to build a multi-surveyor weekly schedule."
        )
    else:
        try:
            team_raw, team_detected_header_row = (
                read_salesforce_or_standard_excel(
                    team_file.getvalue()
                )
            )
            team_upcoming = normalise_upcoming_columns(team_raw)

            if team_detected_header_row > 0:
                st.caption(
                    f"Salesforce report detected: table header found on Excel row "
                    f"{team_detected_header_row + 1}. Report/filter rows above it were ignored."
                )

            if "Postcode" not in team_upcoming.columns:
                st.error(
                    "The master portfolio needs a Postcode column for clustering "
                    "and routing."
                )
            else:
                team_predictions = predict_upcoming(team_upcoming)

                # Both weekly scheduling and strategic planning see the full
                # postcode-bearing portfolio. Work Type / Status no longer block
                # current-week eligibility; drawing/date readiness and a usable
                # duration prediction are the operational gates.
                team_portfolio_input = team_predictions[
                    team_predictions["Postcode"].notna()
                    & (
                        team_predictions["Postcode"]
                        .astype(str)
                        .str.strip()
                        != ""
                    )
                ].copy()

                if team_portfolio_input.empty:
                    st.warning(
                        "There are no portfolio sites with a usable postcode for "
                        "clustering/planning."
                    )
                else:
                    today = datetime.now(LONDON_TZ).date()
                    current_monday = today - timedelta(days=today.weekday())
                    default_team_week = current_monday + timedelta(days=7)

                    def quarter_hour_options(start_hour=6, end_hour=20):
                        options = []
                        for hour in range(start_hour, end_hour + 1):
                            for minute in (0, 15, 30, 45):
                                if hour == end_hour and minute > 0:
                                    continue
                                options.append(time(hour, minute))
                        return options

                    schedule_time_options = quarter_hour_options()
                    time_format = lambda t: t.strftime("%H:%M")

                    t1, t2, t3, t4 = st.columns(4)
                    with t1:
                        team_week_start = st.date_input(
                            "Team week commencing",
                            value=default_team_week,
                            key="team_week_start",
                        )
                    with t2:
                        team_first_survey_clock = st.selectbox(
                            "First survey starts at",
                            schedule_time_options,
                            index=schedule_time_options.index(time(8, 0)),
                            format_func=time_format,
                            key="team_first_survey_clock",
                            help=(
                                "Earliest planned survey start. Home departure is "
                                "back-calculated from Google transit so a long commute "
                                "does not consume the survey window."
                            ),
                        )
                    with t3:
                        team_last_survey_clock = st.selectbox(
                            "Last survey finishes no later than",
                            schedule_time_options,
                            index=schedule_time_options.index(time(15, 0)),
                            format_func=time_format,
                            key="team_last_survey_clock",
                            help=(
                                "Hard deadline for the end of the final survey. "
                                "Travel home happens after this."
                            ),
                        )
                    with t4:
                        team_return_home_clock = st.selectbox(
                            "Return home no later than",
                            schedule_time_options,
                            index=schedule_time_options.index(time(16, 0)),
                            format_func=time_format,
                            key="team_return_home_clock",
                            help=(
                                "Hard home-arrival deadline. The scheduler can finish "
                                "the final survey earlier when the journey home is long."
                            ),
                        )

                    st.caption(
                        "Efficiency rule: the outbound commute is placed before the first "
                        "survey-start target instead of reducing survey time. The final "
                        "survey must satisfy both the survey-finish deadline and the "
                        "return-home deadline."
                    )

                    st.markdown("#### Surveyor availability")
                    st.caption(
                        "Tick the exact dates each person is available in the selected "
                        "week. A surveyor with no dates ticked generates no Google "
                        "routing calls."
                    )

                    selected_week_dates = [
                        team_week_start + timedelta(days=offset)
                        for offset in range(5)
                    ]
                    availability_columns = {
                        d: d.strftime("%a %d %b")
                        for d in selected_week_dates
                    }

                    default_surveyors = pd.DataFrame([
                        {
                            "Name": "Conor Birch",
                            "Start / Finish Location": "Harpenden Station",
                            **{
                                label: True
                                for label in availability_columns.values()
                            },
                        },
                        {
                            "Name": "Rod Harrison",
                            "Start / Finish Location": "Rugby Station",
                            **{
                                label: False
                                for label in availability_columns.values()
                            },
                        },
                        {
                            "Name": "Toby Lawal",
                            "Start / Finish Location": "Chadwell Heath Station",
                            **{
                                label: False
                                for label in availability_columns.values()
                            },
                        },
                        {
                            "Name": "Harrison Grice",
                            "Start / Finish Location": "Gravesend Station",
                            **{
                                label: False
                                for label in availability_columns.values()
                            },
                        },
                        {
                            "Name": "Joe Reynolds",
                            "Start / Finish Location": "Hemel Hempstead Station",
                            **{
                                label: False
                                for label in availability_columns.values()
                            },
                        },
                        {
                            "Name": "Spare Surveyor 1",
                            "Start / Finish Location": "",
                            **{
                                label: False
                                for label in availability_columns.values()
                            },
                        },
                        {
                            "Name": "Spare Surveyor 2",
                            "Start / Finish Location": "",
                            **{
                                label: False
                                for label in availability_columns.values()
                            },
                        },
                    ])

                    availability_config = {
                        label: st.column_config.CheckboxColumn(
                            label,
                            help=f"Available on {d.strftime('%A %d %B %Y')}",
                        )
                        for d, label in availability_columns.items()
                    }

                    edited_surveyors = st.data_editor(
                        default_surveyors,
                        key=(
                            "team_surveyor_editor_"
                            f"{team_week_start.isoformat()}"
                        ),
                        num_rows="fixed",
                        hide_index=True,
                        use_container_width=True,
                        column_config=availability_config,
                    )

                    st.info(
                        "Cost rule: Google route calculations are restricted to "
                        f"{team_week_start.strftime('%d %b %Y')}–"
                        f"{(team_week_start + timedelta(days=6)).strftime('%d %b %Y')}. "
                        "Future weeks influence clustering/drawing priority only."
                    )

                    team_settings = st.columns(4)
                    with team_settings[0]:
                        team_max_candidates = st.number_input(
                            "Max Google candidates per full 5-day surveyor",
                            min_value=15,
                            max_value=100,
                            value=40,
                            step=5,
                            help=(
                                "This is the minimum five-day candidate cap. The app "
                                "automatically raises it when the selected survey window "
                                "and predicted durations show that more jobs could fit, "
                                "so selected working days are not left unused. The whole "
                                "portfolio is still never sent to Google."
                            ),
                        )
                    with team_settings[1]:
                        team_transit_choice = st.selectbox(
                            "Team transit preference",
                            [
                                "Fastest / default",
                                "Less walking",
                                "Fewer transfers",
                            ],
                            key="team_transit_choice",
                        )
                    with team_settings[2]:
                        team_same_postcode = st.number_input(
                            "Minimum close-site transfer (min)",
                            min_value=0,
                            max_value=30,
                            value=5,
                            step=1,
                            key="team_same_postcode",
                            help=(
                                f"Minimum transfer used when buildings share the exact "
                                f"full postcode or are within {NO_GOOGLE_RADIUS_KM:.2f} km "
                                "using Salesforce Latitude/Longitude. Those local site-to-site "
                                "moves bypass Google."
                            ),
                        )
                    with team_settings[3]:
                        use_team_ai_clusters = st.checkbox(
                            "AI cluster selection",
                            value=True,
                            key="team_ai_clusters",
                            help=(
                                "OpenAI sees cheap cluster summaries, not all detailed "
                                "Google route combinations."
                            ),
                        )

                    team_buffers = st.columns(3)
                    with team_buffers[0]:
                        team_travel_leeway = st.number_input(
                            "Team travel leeway (min)",
                            min_value=0,
                            max_value=30,
                            value=5,
                            step=1,
                            key="team_travel_leeway",
                        )
                    with team_buffers[1]:
                        team_pre_buffer = st.number_input(
                            "Before survey (min)",
                            min_value=0,
                            max_value=30,
                            value=5,
                            step=1,
                            key="team_pre_buffer",
                        )
                    with team_buffers[2]:
                        team_post_buffer = st.number_input(
                            "After survey (min)",
                            min_value=0,
                            max_value=30,
                            value=5,
                            step=1,
                            key="team_post_buffer",
                        )

                    st.markdown("#### Weekly notes / special requests")
                    weekly_notes = st.text_area(
                        "Optional requests for this week",
                        value="",
                        height=100,
                        placeholder=(
                            "e.g. Keep Conor as close to Kilburn as possible "
                            "for Thursday"
                        ),
                        help=(
                            "One request per line works best. Notes are tried only "
                            "after a valid baseline schedule is built. If a request "
                            "cannot be satisfied without the existing hard rules, "
                            "the baseline schedule is kept and the note is rejected."
                        ),
                    )
                    special_request_max_minutes = st.number_input(
                        "Maximum distance from requested area (transit minutes)",
                        min_value=5,
                        max_value=90,
                        value=30,
                        step=5,
                        help=(
                            "For a location note, at least one eligible postcode "
                            "cluster must be within this Google transit time of the "
                            "requested area or the note is rejected."
                        ),
                    )
                    st.caption(
                        "Supported in this version: named-surveyor location requests "
                        "for a specific day/date. Notes never override availability, "
                        "drawing eligibility, survey-finish/return-time limits, buffers, duplicate-"
                        "site rules, or the one-week Google horizon. A note may add "
                        "one small area-to-cluster lookup and a trial reroute for the "
                        "affected surveyor, but never routes a future week."
                    )

                    generate_team_ai_summary = st.checkbox(
                        "Generate one AI summary of the final team week",
                        value=True,
                        key="team_ai_summary",
                    )

                    team_google_key = get_secret("GOOGLE_MAPS_API_KEY")
                    team_openai_key = get_secret("OPENAI_API_KEY")
                    team_openai_model = get_secret("OPENAI_MODEL", "gpt-5.6")
                    team_tfl_key = get_secret("TFL_API_KEY")
                    team_met_key = get_secret("MET_OFFICE_API_KEY")
                    team_met_endpoint = get_secret(
                        "MET_OFFICE_GLOBAL_SPOT_URL"
                    )

                    status_cols = st.columns(4)
                    status_cols[0].metric(
                        "Portfolio rows",
                        len(team_predictions),
                    )
                    status_cols[1].metric(
                        "Rows considered strategically",
                        len(team_portfolio_input),
                    )
                    status_cols[2].metric(
                        "Google Routes",
                        "Ready" if team_google_key else "Missing",
                    )
                    status_cols[3].metric(
                        "OpenAI",
                        "Ready" if team_openai_key else "Missing",
                    )

                    if st.button(
                        "Create weekly schedules",
                        type="primary",
                        key="create_team_week",
                        disabled=not bool(team_google_key.strip()),
                    ):
                        if team_week_start.weekday() != 0:
                            st.error(
                                "Week commencing must be a Monday."
                            )
                        elif team_last_survey_clock <= team_first_survey_clock:
                            st.error(
                                "Last survey finish must be after the first survey start."
                            )
                        elif team_return_home_clock <= team_last_survey_clock:
                            st.error(
                                "Return-home deadline must be after the last-survey "
                                "finish deadline."
                            )
                        else:
                            active_surveyors = []
                            invalid_active_rows = []

                            for _, row in edited_surveyors.iterrows():
                                name = str(row.get("Name", "")).strip()
                                location = str(
                                    row.get(
                                        "Start / Finish Location",
                                        "",
                                    )
                                ).strip()

                                available_dates = [
                                    d
                                    for d, label in availability_columns.items()
                                    if bool(row.get(label, False))
                                ]

                                # No dates ticked = not working this week.
                                if not available_dates:
                                    continue

                                if not name:
                                    invalid_active_rows.append(
                                        "A surveyor with availability selected "
                                        "has no name."
                                    )
                                    continue
                                if not location:
                                    invalid_active_rows.append(
                                        f"{name} has availability selected but "
                                        "no start / finish location."
                                    )
                                    continue

                                active_surveyors.append(
                                    SurveyorConfig(
                                        name=name,
                                        start_location=location,
                                        available_dates=available_dates,
                                    )
                                )

                            all_available_dates = sorted({
                                d
                                for surveyor in active_surveyors
                                for d in (surveyor.available_dates or [])
                            })

                            if invalid_active_rows:
                                for message in invalid_active_rows:
                                    st.error(message)
                            elif not active_surveyors:
                                st.error(
                                    "No surveyors have any availability selected for this week."
                                )
                            else:
                                try:
                                    with st.spinner(
                                        "Clustering the portfolio, allocating work "
                                        "across the team, then routing only each "
                                        "person's shortlist..."
                                    ):
                                        team_portfolio = add_portfolio_fields(
                                            team_portfolio_input,
                                            target_week_start=team_week_start,
                                            today=today,
                                        )
                                        team_eligible = team_portfolio[
                                            team_portfolio[
                                                "Eligible for Selected Week"
                                            ] == True
                                        ].copy()

                                        coordinate_count = int(
                                            team_portfolio.get(
                                                "Coordinate Available",
                                                pd.Series(False, index=team_portfolio.index),
                                            ).astype(bool).sum()
                                        )
                                        st.caption(
                                            f"Coordinate clustering active for {coordinate_count:,} of "
                                            f"{len(team_portfolio):,} portfolio sites. Strategic clusters "
                                            f"use a {GEOGRAPHIC_CLUSTER_MAX_DIAMETER_KM:.1f} km maximum "
                                            f"diameter; site-to-site Google routing is bypassed within "
                                            f"{NO_GOOGLE_RADIUS_KM:.2f} km. Rows without coordinates fall "
                                            "back to postcode-district clustering."
                                        )

                                        if team_eligible.empty:
                                            raise ValueError(
                                                "No sites are eligible for the "
                                                "selected team week."
                                            )

                                        team_cluster_summary = (
                                            build_cluster_summary(
                                                team_portfolio,
                                                team_week_start,
                                            )
                                        )

                                        team_period_end = (
                                            team_week_start
                                            + timedelta(days=6)
                                        )

                                        team_tfl_summary = (
                                            "TfL integration not configured."
                                        )
                                        if team_tfl_key:
                                            team_tfl_summary = TfLClient(
                                                team_tfl_key
                                            ).disruption_summary(
                                                team_week_start,
                                                team_period_end,
                                            )

                                        team_weather_summary = (
                                            "Met Office integration not configured."
                                        )
                                        if (
                                            team_met_key
                                            and team_met_endpoint
                                        ):
                                            team_weather_summary = (
                                                MetOfficeClient(
                                                    team_met_key,
                                                    team_met_endpoint,
                                                ).forecast_summary(
                                                    latitude=51.65,
                                                    longitude=-0.20,
                                                    start_date=team_week_start,
                                                    end_date=team_period_end,
                                                )
                                            )

                                        # Time-aware candidate capacity.
                                        #
                                        # The old logic scaled a fixed 40-site five-day
                                        # cap by availability (e.g. 3 days -> 24 sites).
                                        # That can under-fill later selected days when
                                        # surveys are short/dense: all 24 candidates may
                                        # be completed in the first two days.
                                        #
                                        # Keep the user's existing cap as a MINIMUM, but
                                        # automatically expand it using:
                                        #   - selected first/last survey times;
                                        #   - the lower-quartile real prediction duration;
                                        #   - existing pre/post buffers;
                                        #   - the close-site transfer assumption;
                                        #   - a 50% choice reserve so Google has enough
                                        #     alternatives after feasibility/endgame rules.
                                        window_start_dt = datetime.combine(
                                            team_week_start,
                                            team_first_survey_clock,
                                        )
                                        window_end_dt = datetime.combine(
                                            team_week_start,
                                            team_last_survey_clock,
                                        )
                                        survey_window_minutes = max(
                                            60.0,
                                            (
                                                window_end_dt
                                                - window_start_dt
                                            ).total_seconds()
                                            / 60.0,
                                        )

                                        # Lunch is a protected 30-minute break whenever
                                        # the selected survey window spans the lunch period.
                                        lunch_overlap = (
                                            team_first_survey_clock
                                            <= time(13, 0)
                                            and team_last_survey_clock
                                            >= time(11, 45)
                                        )
                                        usable_survey_window_minutes = max(
                                            60.0,
                                            survey_window_minutes
                                            - (30.0 if lunch_overlap else 0.0),
                                        )

                                        eligible_duration_values = pd.to_numeric(
                                            team_portfolio.loc[
                                                team_portfolio[
                                                    "Eligible for Selected Week"
                                                ]
                                                == True,
                                                "Planning Duration (Minutes)",
                                            ],
                                            errors="coerce",
                                        )
                                        eligible_duration_values = (
                                            eligible_duration_values[
                                                eligible_duration_values > 0
                                            ]
                                            .dropna()
                                        )

                                        if len(eligible_duration_values) >= 4:
                                            reference_survey_minutes = float(
                                                eligible_duration_values.quantile(
                                                    0.25
                                                )
                                            )
                                        elif len(eligible_duration_values) > 0:
                                            reference_survey_minutes = float(
                                                eligible_duration_values.median()
                                            )
                                        else:
                                            reference_survey_minutes = 30.0

                                        # Do not let one exceptionally tiny prediction
                                        # explode Google cost.
                                        reference_survey_minutes = max(
                                            10.0,
                                            reference_survey_minutes,
                                        )
                                        assumed_local_transfer_minutes = max(
                                            3.0,
                                            min(
                                                10.0,
                                                float(team_same_postcode),
                                            ),
                                        )
                                        estimated_cycle_minutes = max(
                                            15.0,
                                            reference_survey_minutes
                                            + float(team_pre_buffer)
                                            + float(team_post_buffer)
                                            + assumed_local_transfer_minutes,
                                        )

                                        estimated_jobs_per_day = max(
                                            1,
                                            math.ceil(
                                                usable_survey_window_minutes
                                                / estimated_cycle_minutes
                                            ),
                                        )

                                        # Extra candidate choice is deliberate: some
                                        # candidates will fail return-home feasibility,
                                        # be held as endgame anchors, or be inferior to
                                        # denser alternatives. Cap at 25 candidates/day
                                        # to retain the existing cost-control philosophy.
                                        time_aware_candidates_per_day = min(
                                            25,
                                            max(
                                                8,
                                                math.ceil(
                                                    estimated_jobs_per_day
                                                    * 1.50
                                                ),
                                            ),
                                        )
                                        time_aware_full_week_cap = (
                                            time_aware_candidates_per_day * 5
                                        )
                                        effective_full_week_candidate_cap = max(
                                            int(team_max_candidates),
                                            int(time_aware_full_week_cap),
                                        )

                                        effective_candidate_caps = {
                                            surveyor.name: max(
                                                1,
                                                round(
                                                    effective_full_week_candidate_cap
                                                    * min(
                                                        5,
                                                        len(
                                                            surveyor.available_dates
                                                            or []
                                                        ),
                                                    )
                                                    / 5
                                                ),
                                            )
                                            for surveyor in active_surveyors
                                        }
                                        total_team_candidate_cap = sum(
                                            effective_candidate_caps.values()
                                        )

                                        team_planner = None
                                        if (
                                            use_team_ai_clusters
                                            or generate_team_ai_summary
                                            or bool(str(weekly_notes).strip())
                                        ):
                                            if not team_openai_key:
                                                raise ValueError(
                                                    "OPENAI_API_KEY is missing "
                                                    "from secrets."
                                                )
                                            team_planner = (
                                                OpenAISchedulePlanner(
                                                    api_key=team_openai_key,
                                                    model=team_openai_model,
                                                )
                                            )

                                        if use_team_ai_clusters:
                                            (
                                                team_cluster_decisions,
                                                team_deferred_clusters,
                                                team_cluster_strategy,
                                            ) = team_planner.select_clusters(
                                                cluster_summary=(
                                                    team_cluster_summary
                                                    .to_dict(
                                                        orient="records"
                                                    )
                                                ),
                                                week_label=(
                                                    f"{team_week_start} to "
                                                    f"{team_period_end}"
                                                ),
                                                working_days=len(
                                                    all_available_dates
                                                ),
                                                max_sites_for_google=(
                                                    total_team_candidate_cap
                                                ),
                                                tfl_summary=(
                                                    team_tfl_summary
                                                ),
                                                weather_summary=(
                                                    team_weather_summary
                                                ),
                                                team_size=len(
                                                    active_surveyors
                                                ),
                                            )
                                            team_cluster_choices = [
                                                {
                                                    "cluster": d.cluster,
                                                    "priority": d.priority,
                                                    "target_sites": (
                                                        d.target_sites
                                                    ),
                                                    "decision": d.decision,
                                                    "reason": d.reason,
                                                }
                                                for d in (
                                                    team_cluster_decisions
                                                )
                                            ]
                                        else:
                                            team_cluster_choices = (
                                                deterministic_cluster_choices(
                                                    team_cluster_summary,
                                                    total_team_candidate_cap,
                                                )
                                            )
                                            team_deferred_clusters = []
                                            team_cluster_strategy = (
                                                "AI cluster selection disabled. "
                                                "The largest eligible postcode "
                                                "clusters were used."
                                            )

                                        if not team_cluster_choices:
                                            raise ValueError(
                                                "No clusters were selected for "
                                                "the team week."
                                            )

                                        # Rolling-horizon endgame guardrail. This is a
                                        # cheap portfolio-only adjustment: it can leave
                                        # a small number of released sites as future
                                        # geographic anchors, replace them with
                                        # non-anchor candidates elsewhere, and only
                                        # release anchors if current-week candidate
                                        # capacity would otherwise be lost. No Google
                                        # calls are made here.
                                        (
                                            team_cluster_choices,
                                            team_endgame_plan_df,
                                        ) = endgame_adjust_cluster_choices(
                                            cluster_choices=team_cluster_choices,
                                            cluster_summary=team_cluster_summary,
                                            max_sites_for_google=(
                                                total_team_candidate_cap
                                            ),
                                        )

                                        if team_endgame_plan_df.empty:
                                            team_endgame_plan_df = pd.DataFrame(
                                                columns=[
                                                    "Cluster",
                                                    "Eligible Now",
                                                    "Future Pipeline Sites",
                                                    "Future Plan Drafting",
                                                    "Future Under Preparation",
                                                    "Future Work Done",
                                                    "Future Plan Drafting Work Done",
                                                    "Future Parent Work Done",
                                                    "Endgame Risk",
                                                    "Suggested Anchor Reserve",
                                                    "Candidate Target This Week",
                                                    "Eligible Sites Left Outside Shortlist",
                                                    "Endgame Reason",
                                                ]
                                            )

                                        drawing_priority_queue = (
                                            build_drawing_priority_queue(
                                                portfolio=team_portfolio,
                                                cluster_summary=team_cluster_summary,
                                                selected_cluster_choices=(
                                                    team_cluster_choices
                                                ),
                                                target_week_start=(
                                                    team_week_start
                                                ),
                                            )
                                        )

                                        team_router_pref = {
                                            "Fastest / default": None,
                                            "Less walking": "LESS_WALKING",
                                            "Fewer transfers": (
                                                "FEWER_TRANSFERS"
                                            ),
                                        }

                                        # Google enters here for the first time:
                                        # tiny surveyor-home -> selected-cluster matrix.
                                        team_router = GoogleTransitRouter(
                                            api_key=team_google_key,
                                            transit_preference=(
                                                team_router_pref[
                                                    team_transit_choice
                                                ]
                                            ),
                                        )

                                        team_representatives = (
                                            representative_sites(
                                                team_portfolio,
                                                team_cluster_choices,
                                                team_week_start,
                                            )
                                        )

                                        # The home->cluster allocation matrix is only a
                                        # cheap comparative signal. Probe shortly before
                                        # the first-survey target; the detailed scheduler
                                        # later back-calculates each person's real departure.
                                        allocation_departure = (
                                            datetime.combine(
                                                min(all_available_dates),
                                                team_first_survey_clock,
                                                tzinfo=LONDON_TZ,
                                            )
                                            - timedelta(minutes=90)
                                        )

                                        team_home_cluster_matrix = (
                                            home_to_cluster_matrix(
                                                team_router,
                                                active_surveyors,
                                                team_representatives,
                                                allocation_departure,
                                            )
                                        )

                                        team_allocations = (
                                            allocate_cluster_targets(
                                                cluster_choices=(
                                                    team_cluster_choices
                                                ),
                                                cluster_summary=(
                                                    team_cluster_summary
                                                ),
                                                surveyors=active_surveyors,
                                                travel_matrix=(
                                                    team_home_cluster_matrix
                                                ),
                                                max_sites_per_surveyor=int(
                                                    effective_full_week_candidate_cap
                                                ),
                                            )
                                        )

                                        if not team_allocations:
                                            raise ValueError(
                                                "Google could not create any "
                                                "usable surveyor-to-cluster "
                                                "allocations."
                                            )

                                        team_shortlists = (
                                            build_team_shortlists(
                                                portfolio=team_portfolio,
                                                allocations=team_allocations,
                                                surveyors=active_surveyors,
                                                target_week_start=(
                                                    team_week_start
                                                ),
                                                max_sites_per_surveyor=int(
                                                    effective_full_week_candidate_cap
                                                ),
                                            )
                                        )

                                        team_results = {}
                                        combined_candidate_frames = []

                                        for surveyor in active_surveyors:
                                            surveyor_df = team_shortlists.get(
                                                surveyor.name,
                                                pd.DataFrame(),
                                            )
                                            if surveyor_df.empty:
                                                team_results[
                                                    surveyor.name
                                                ] = None
                                                continue

                                            combined_candidate_frames.append(
                                                surveyor_df
                                            )

                                            surveyor_sites = (
                                                site_dataframe_to_dicts(
                                                    surveyor_df
                                                )
                                            )

                                            surveyor_scheduler = (
                                                DailyTransitScheduler(
                                                    router=team_router,
                                                    home_location=(
                                                        surveyor.start_location
                                                    ),
                                                    same_postcode_transfer_minutes=int(
                                                        team_same_postcode
                                                    ),
                                                    travel_leeway_minutes=int(
                                                        team_travel_leeway
                                                    ),
                                                    pre_survey_buffer_minutes=int(
                                                        team_pre_buffer
                                                    ),
                                                    post_survey_buffer_minutes=int(
                                                        team_post_buffer
                                                    ),
                                                    ai_priority_weight_minutes=(
                                                        15.0
                                                    ),
                                                )
                                            )

                                            surveyor_dates = list(
                                                surveyor.available_dates or []
                                            )

                                            if not surveyor_dates:
                                                team_results[
                                                    surveyor.name
                                                ] = None
                                                continue

                                            team_results[
                                                surveyor.name
                                            ] = (
                                                surveyor_scheduler.build_week(
                                                    sites=surveyor_sites,
                                                    dates=surveyor_dates,
                                                    first_survey_start_clock=(
                                                        team_first_survey_clock
                                                    ),
                                                    latest_survey_finish_clock=(
                                                        team_last_survey_clock
                                                    ),
                                                    latest_return_clock=(
                                                        team_return_home_clock
                                                    ),
                                                    timezone=LONDON_TZ,
                                                )
                                            )

                                        # Optional weekly notes are transactional:
                                        # the baseline schedule above remains the fallback.
                                        # A note only replaces an affected surveyor's week
                                        # if the trial schedule can actually satisfy it
                                        # while retaining every hard scheduler constraint.
                                        special_request_results = []

                                        if str(weekly_notes or "").strip():
                                            parsed_requests = (
                                                team_planner.parse_week_notes(
                                                    notes_text=weekly_notes,
                                                    surveyors=[
                                                        {
                                                            "name": s.name,
                                                            "available_dates": [
                                                                d.isoformat()
                                                                for d in (
                                                                    s.available_dates
                                                                    or []
                                                                )
                                                            ],
                                                        }
                                                        for s in active_surveyors
                                                    ],
                                                    week_dates=[
                                                        d.isoformat()
                                                        for d in selected_week_dates
                                                    ],
                                                )
                                            )

                                            request_representatives = (
                                                all_cluster_representatives(
                                                    team_portfolio,
                                                    team_week_start,
                                                )
                                            )

                                            surveyor_lookup = {
                                                s.name.lower(): s
                                                for s in active_surveyors
                                            }

                                            for request in parsed_requests:
                                                if (
                                                    not request.supported
                                                    or request.request_type
                                                    != "location_preference"
                                                ):
                                                    special_request_results.append(
                                                        SpecialRequestResult(
                                                            raw_note=request.raw_note,
                                                            surveyor_name=(
                                                                request.surveyor_name
                                                            ),
                                                            requested_date=(
                                                                request.requested_date
                                                            ),
                                                            location=request.location,
                                                            status="Rejected",
                                                            reason=(
                                                                request.rejection_reason
                                                                or "Unsupported or ambiguous weekly note."
                                                            ),
                                                        )
                                                    )
                                                    continue

                                                surveyor = surveyor_lookup.get(
                                                    request.surveyor_name.lower()
                                                )
                                                if surveyor is None:
                                                    special_request_results.append(
                                                        SpecialRequestResult(
                                                            raw_note=request.raw_note,
                                                            surveyor_name=(
                                                                request.surveyor_name
                                                            ),
                                                            requested_date=(
                                                                request.requested_date
                                                            ),
                                                            location=request.location,
                                                            status="Rejected",
                                                            reason=(
                                                                "The named surveyor is not active "
                                                                "for the selected week."
                                                            ),
                                                        )
                                                    )
                                                    continue

                                                requested_ts = pd.to_datetime(
                                                    request.requested_date,
                                                    errors="coerce",
                                                )
                                                if pd.isna(requested_ts):
                                                    special_request_results.append(
                                                        SpecialRequestResult(
                                                            raw_note=request.raw_note,
                                                            surveyor_name=surveyor.name,
                                                            requested_date=(
                                                                request.requested_date
                                                            ),
                                                            location=request.location,
                                                            status="Rejected",
                                                            reason=(
                                                                "The requested day/date could not "
                                                                "be resolved inside the selected week."
                                                            ),
                                                        )
                                                    )
                                                    continue

                                                requested_date = requested_ts.date()
                                                if requested_date not in (
                                                    surveyor.available_dates or []
                                                ):
                                                    special_request_results.append(
                                                        SpecialRequestResult(
                                                            raw_note=request.raw_note,
                                                            surveyor_name=surveyor.name,
                                                            requested_date=(
                                                                requested_date.isoformat()
                                                            ),
                                                            location=request.location,
                                                            status="Rejected",
                                                            reason=(
                                                                f"{surveyor.name} is not marked "
                                                                "available on that date."
                                                            ),
                                                        )
                                                    )
                                                    continue

                                                if not request.location.strip():
                                                    special_request_results.append(
                                                        SpecialRequestResult(
                                                            raw_note=request.raw_note,
                                                            surveyor_name=surveyor.name,
                                                            requested_date=(
                                                                requested_date.isoformat()
                                                            ),
                                                            location=request.location,
                                                            status="Rejected",
                                                            reason=(
                                                                "No location could be identified "
                                                                "from the note."
                                                            ),
                                                        )
                                                    )
                                                    continue

                                                (
                                                    nearby_cluster,
                                                    nearest_minutes,
                                                    _request_ranked_clusters,
                                                ) = choose_nearby_cluster(
                                                    router=team_router,
                                                    request_location=(
                                                        request.location
                                                    ),
                                                    requested_date=requested_date,
                                                    start_clock=team_first_survey_clock,
                                                    timezone=LONDON_TZ,
                                                    representatives=(
                                                        request_representatives
                                                    ),
                                                    max_minutes=float(
                                                        special_request_max_minutes
                                                    ),
                                                )

                                                if nearby_cluster is None:
                                                    if nearest_minutes is None:
                                                        reason = (
                                                            "Google could not identify a reachable "
                                                            "eligible cluster near the requested area."
                                                        )
                                                    else:
                                                        reason = (
                                                            f"The nearest eligible cluster was about "
                                                            f"{nearest_minutes:.0f} minutes from "
                                                            f"{request.location}, above the "
                                                            f"{int(special_request_max_minutes)}-minute "
                                                            "request limit."
                                                        )
                                                    special_request_results.append(
                                                        SpecialRequestResult(
                                                            raw_note=request.raw_note,
                                                            surveyor_name=surveyor.name,
                                                            requested_date=(
                                                                requested_date.isoformat()
                                                            ),
                                                            location=request.location,
                                                            status="Rejected",
                                                            anchor_to_cluster_minutes=(
                                                                nearest_minutes
                                                            ),
                                                            reason=reason,
                                                        )
                                                    )
                                                    continue

                                                target_cluster = str(
                                                    nearby_cluster["cluster"]
                                                )
                                                excluded_refs = (
                                                    scheduled_reference_set(
                                                        team_results,
                                                        exclude_name=surveyor.name,
                                                    )
                                                )

                                                cap = int(
                                                    effective_candidate_caps.get(
                                                        surveyor.name,
                                                        team_max_candidates,
                                                    )
                                                )
                                                day_count = max(
                                                    1,
                                                    len(
                                                        surveyor.available_dates
                                                        or []
                                                    ),
                                                )
                                                target_count = max(
                                                    4,
                                                    (cap + day_count - 1)
                                                    // day_count,
                                                )

                                                current_shortlist = (
                                                    team_shortlists.get(
                                                        surveyor.name,
                                                        pd.DataFrame(),
                                                    )
                                                )
                                                trial_df = build_trial_dataframe(
                                                    portfolio=team_portfolio,
                                                    current_shortlist=(
                                                        current_shortlist
                                                    ),
                                                    target_cluster=target_cluster,
                                                    surveyor_name=surveyor.name,
                                                    requested_date=requested_date,
                                                    raw_note=request.raw_note,
                                                    excluded_refs=excluded_refs,
                                                    candidate_cap=cap,
                                                    target_candidate_count=(
                                                        target_count
                                                    ),
                                                    target_week_start=(
                                                        team_week_start
                                                    ),
                                                )

                                                if trial_df.empty:
                                                    special_request_results.append(
                                                        SpecialRequestResult(
                                                            raw_note=request.raw_note,
                                                            surveyor_name=surveyor.name,
                                                            requested_date=(
                                                                requested_date.isoformat()
                                                            ),
                                                            location=request.location,
                                                            status="Rejected",
                                                            target_cluster=target_cluster,
                                                            anchor_to_cluster_minutes=(
                                                                nearest_minutes
                                                            ),
                                                            reason=(
                                                                "No unallocated eligible sites were "
                                                                "available in the nearby cluster."
                                                            ),
                                                        )
                                                    )
                                                    continue

                                                trial_scheduler = DailyTransitScheduler(
                                                    router=team_router,
                                                    home_location=(
                                                        surveyor.start_location
                                                    ),
                                                    same_postcode_transfer_minutes=int(
                                                        team_same_postcode
                                                    ),
                                                    travel_leeway_minutes=int(
                                                        team_travel_leeway
                                                    ),
                                                    pre_survey_buffer_minutes=int(
                                                        team_pre_buffer
                                                    ),
                                                    post_survey_buffer_minutes=int(
                                                        team_post_buffer
                                                    ),
                                                    ai_priority_weight_minutes=15.0,
                                                )
                                                trial_result = (
                                                    trial_scheduler.build_week(
                                                        sites=(
                                                            site_dataframe_to_dicts(
                                                                trial_df
                                                            )
                                                        ),
                                                        dates=list(
                                                            surveyor.available_dates
                                                            or []
                                                        ),
                                                        first_survey_start_clock=(
                                                            team_first_survey_clock
                                                        ),
                                                        latest_survey_finish_clock=(
                                                            team_last_survey_clock
                                                        ),
                                                        latest_return_clock=(
                                                            team_return_home_clock
                                                        ),
                                                        timezone=LONDON_TZ,
                                                    )
                                                )

                                                if result_uses_cluster_on_date(
                                                    trial_result,
                                                    requested_date,
                                                    target_cluster,
                                                ):
                                                    # Transaction commits only here.
                                                    team_results[
                                                        surveyor.name
                                                    ] = trial_result
                                                    team_shortlists[
                                                        surveyor.name
                                                    ] = trial_df
                                                    special_request_results.append(
                                                        SpecialRequestResult(
                                                            raw_note=request.raw_note,
                                                            surveyor_name=surveyor.name,
                                                            requested_date=(
                                                                requested_date.isoformat()
                                                            ),
                                                            location=request.location,
                                                            status="Accepted",
                                                            target_cluster=target_cluster,
                                                            anchor_to_cluster_minutes=(
                                                                nearest_minutes
                                                            ),
                                                            reason=(
                                                                f"A valid trial schedule kept "
                                                                f"{surveyor.name} in/around "
                                                                f"{target_cluster} on "
                                                                f"{requested_date.strftime('%A')}. "
                                                                "All hard scheduling rules remained valid."
                                                            ),
                                                        )
                                                    )
                                                else:
                                                    special_request_results.append(
                                                        SpecialRequestResult(
                                                            raw_note=request.raw_note,
                                                            surveyor_name=surveyor.name,
                                                            requested_date=(
                                                                requested_date.isoformat()
                                                            ),
                                                            location=request.location,
                                                            status="Rejected",
                                                            target_cluster=target_cluster,
                                                            anchor_to_cluster_minutes=(
                                                                nearest_minutes
                                                            ),
                                                            reason=(
                                                                "The request could not be fitted on the "
                                                                "requested date while preserving the "
                                                                "existing return-time and scheduling rules. "
                                                                "The baseline schedule was kept."
                                                            ),
                                                        )
                                                    )

                                        team_allocations_df = (
                                            allocations_dataframe(
                                                team_allocations
                                            )
                                        )

                                        final_shortlist_frames = [
                                            df
                                            for df in team_shortlists.values()
                                            if df is not None and not df.empty
                                        ]
                                        if final_shortlist_frames:
                                            team_google_shortlist = pd.concat(
                                                final_shortlist_frames,
                                                ignore_index=True,
                                            ).drop_duplicates()
                                        else:
                                            team_google_shortlist = (
                                                team_eligible.head(0)
                                            )

                                        special_request_results_df = (
                                            request_results_dataframe(
                                                special_request_results
                                            )
                                        )

                                        # Build combined team outputs.
                                        team_summary_rows = []
                                        combined_schedule_frames = []

                                        for surveyor in active_surveyors:
                                            result = team_results.get(
                                                surveyor.name
                                            )
                                            if result is None:
                                                team_summary_rows.append({
                                                    "Surveyor": (
                                                        surveyor.name
                                                    ),
                                                    "Start / Finish": (
                                                        surveyor.start_location
                                                    ),
                                                    "Available Dates": ", ".join(
                                                        d.strftime("%a %d %b")
                                                        for d in (
                                                            surveyor.available_dates
                                                            or []
                                                        )
                                                    ),
                                                    "Google Candidate Cap": (
                                                        effective_candidate_caps.get(
                                                            surveyor.name, 0
                                                        )
                                                    ),
                                                    "Surveys": 0,
                                                    "Survey Minutes": 0,
                                                    "Travel Minutes": 0,
                                                    "Shortlist Sites": 0,
                                                })
                                                continue

                                            team_summary_rows.append({
                                                "Surveyor": surveyor.name,
                                                "Start / Finish": (
                                                    surveyor.start_location
                                                ),
                                                "Available Dates": ", ".join(
                                                    d.strftime("%a %d %b")
                                                    for d in (
                                                        surveyor.available_dates
                                                        or []
                                                    )
                                                ),
                                                "Google Candidate Cap": (
                                                    effective_candidate_caps.get(
                                                        surveyor.name, 0
                                                    )
                                                ),
                                                "Surveys": (
                                                    result.total_surveys
                                                ),
                                                "Survey Minutes": (
                                                    result.total_survey_minutes
                                                ),
                                                "Travel Minutes": (
                                                    result.total_travel_minutes
                                                ),
                                                "Shortlist Sites": len(
                                                    team_shortlists.get(
                                                        surveyor.name,
                                                        pd.DataFrame(),
                                                    )
                                                ),
                                            })

                                            full_df = (
                                                result
                                                .full_schedule_dataframe()
                                                .copy()
                                            )
                                            if not full_df.empty:
                                                full_df.insert(
                                                    0,
                                                    "Surveyor",
                                                    surveyor.name,
                                                )
                                                combined_schedule_frames.append(
                                                    full_df
                                                )

                                        team_summary_df = pd.DataFrame(
                                            team_summary_rows
                                        )
                                        if combined_schedule_frames:
                                            combined_team_schedule = (
                                                pd.concat(
                                                    combined_schedule_frames,
                                                    ignore_index=True,
                                                )
                                            )
                                        else:
                                            combined_team_schedule = (
                                                pd.DataFrame()
                                            )

                                        salesforce_copy_df = build_salesforce_copy(
                                            combined_team_schedule,
                                            team_upcoming,
                                        )

                                        cluster_matrix_rows = []
                                        for surveyor in active_surveyors:
                                            for rep in (
                                                team_representatives
                                            ):
                                                cluster_matrix_rows.append({
                                                    "Surveyor": (
                                                        surveyor.name
                                                    ),
                                                    "Cluster": (
                                                        rep["cluster"]
                                                    ),
                                                    "Home to Cluster "
                                                    "(Minutes)": (
                                                        team_home_cluster_matrix
                                                        .get(
                                                            (
                                                                surveyor.name,
                                                                rep[
                                                                    "cluster"
                                                                ],
                                                            )
                                                        )
                                                    ),
                                                })
                                        team_home_cluster_df = pd.DataFrame(
                                            cluster_matrix_rows
                                        )

                                        team_strategy_context = (
                                            "Strategic cluster selection:\n"
                                            f"{team_cluster_strategy}\n\n"
                                            "Endgame / orphan-risk plan:\n"
                                            f"{team_endgame_plan_df.to_dict(orient='records')}\n\n"
                                            "Team cluster allocations:\n"
                                            f"{team_allocations_df.to_dict(orient='records')}\n\n"
                                            "Weekly special-request results:\n"
                                            f"{special_request_results_df.to_dict(orient='records')}"
                                        )

                                    st.markdown("#### Team portfolio pre-filter")
                                    tm1, tm2, tm3, tm4 = st.columns(4)
                                    tm1.metric(
                                        "Active surveyors",
                                        len(active_surveyors),
                                    )
                                    tm2.metric(
                                        "Eligible portfolio",
                                        len(team_eligible),
                                    )
                                    tm3.metric(
                                        "Detailed shortlist sites",
                                        len(team_google_shortlist),
                                    )
                                    reduction = (
                                        100
                                        * (
                                            len(team_eligible)
                                            - len(team_google_shortlist)
                                        )
                                        / len(team_eligible)
                                        if len(team_eligible)
                                        else 0
                                    )
                                    tm4.metric(
                                        "Filtered before routing",
                                        f"{reduction:.0f}%",
                                    )

                                    st.caption(
                                        "The only cross-team Google comparison is "
                                        f"{len(active_surveyors)} surveyor homes × "
                                        f"{len(team_representatives)} selected cluster "
                                        "representatives, followed by each person's "
                                        "own capped shortlist."
                                    )

                                    # Reporting only: show the actual Google usage
                                    # plus the site-to-site calculations avoided by
                                    # coordinate/postcode proximity rules.
                                    google_usage = (
                                        team_router.get_usage_stats()
                                        if hasattr(
                                            team_router,
                                            "get_usage_stats",
                                        )
                                        else {}
                                    )

                                    schedule_routing_stats = {
                                        "google_matrix_destinations": 0,
                                        "same_postcode_bypasses": 0,
                                        "coordinate_radius_bypasses": 0,
                                        "legacy_campus_bypasses": 0,
                                        "collapsed_nearby_destinations": 0,
                                        "routing_calculations_avoided": 0,
                                    }
                                    for _result in team_results.values():
                                        if _result is None:
                                            continue
                                        _stats = getattr(
                                            _result,
                                            "routing_stats",
                                            {},
                                        ) or {}
                                        for _key in schedule_routing_stats:
                                            schedule_routing_stats[_key] += int(
                                                _stats.get(_key, 0) or 0
                                            )

                                    _avoided = int(
                                        schedule_routing_stats[
                                            "routing_calculations_avoided"
                                        ]
                                    )
                                    _site_matrix_sent = int(
                                        schedule_routing_stats[
                                            "google_matrix_destinations"
                                        ]
                                    )
                                    _site_matrix_total = (
                                        _site_matrix_sent + _avoided
                                    )
                                    _avoidance_pct = (
                                        100.0
                                        * _avoided
                                        / _site_matrix_total
                                        if _site_matrix_total
                                        else 0.0
                                    )

                                    st.markdown("#### Google routing usage")
                                    gu1, gu2, gu3, gu4 = st.columns(4)
                                    gu1.metric(
                                        "Unique Google matrix destinations",
                                        int(
                                            google_usage.get(
                                                "unique_matrix_destinations",
                                                0,
                                            )
                                        ),
                                        help=(
                                            "Unique destination strings actually "
                                            "sent to Google's Route Matrix during "
                                            "this planning run. This includes the "
                                            "small home-to-cluster comparison as "
                                            "well as site-routing destinations."
                                        ),
                                    )
                                    gu2.metric(
                                        "Google matrix calculations",
                                        int(
                                            google_usage.get(
                                                "matrix_destination_elements",
                                                0,
                                            )
                                        ),
                                        help=(
                                            "Total destination elements actually "
                                            "sent to Google Route Matrix. The same "
                                            "site can appear more than once as the "
                                            "route is re-evaluated through the day."
                                        ),
                                    )
                                    gu3.metric(
                                        "Google API requests",
                                        int(
                                            google_usage.get(
                                                "total_google_http_requests",
                                                0,
                                            )
                                        ),
                                        help=(
                                            "Actual Compute Routes + Compute Route "
                                            "Matrix HTTP requests made in this run."
                                        ),
                                    )
                                    gu4.metric(
                                        "Local routing avoided",
                                        _avoided,
                                        delta=f"{_avoidance_pct:.0f}% site-to-site",
                                        help=(
                                            "Site-to-site destination calculations "
                                            "avoided by same-postcode, coordinate "
                                            "proximity and nearby-destination "
                                            "collapse rules."
                                        ),
                                    )

                                    st.caption(
                                        "Site-to-site savings breakdown: "
                                        f"{schedule_routing_stats['coordinate_radius_bypasses']} "
                                        f"within {NO_GOOGLE_RADIUS_KM:.2f} km, "
                                        f"{schedule_routing_stats['same_postcode_bypasses']} "
                                        "same-full-postcode bypasses, "
                                        f"{schedule_routing_stats['legacy_campus_bypasses']} "
                                        "legacy same-campus bypasses, and "
                                        f"{schedule_routing_stats['collapsed_nearby_destinations']} "
                                        "nearby destinations collapsed before Google. "
                                        "The Google API figures also include home-to-cluster "
                                        "comparison and first-site/return-home feasibility "
                                        "routing, so they intentionally measure actual API "
                                        "usage rather than only final scheduled buildings."
                                    )

                                    st.markdown("#### Team cluster strategy")
                                    st.info(team_cluster_strategy)

                                    st.markdown("#### Endgame / orphan-risk planning")
                                    st.caption(
                                        "This looks across the remaining portfolio before "
                                        "Google routing. Any currently eligible site can be "
                                        "left outside this week's shortlist as a future "
                                        "geographic anchor when not-yet-eligible work remains "
                                        "nearby. Anchors are restored if they are needed to "
                                        "avoid under-supplying the current week."
                                    )
                                    if team_endgame_plan_df.empty:
                                        st.caption(
                                            "No endgame planning signals were available for "
                                            "the current portfolio."
                                        )
                                    else:
                                        st.dataframe(
                                            team_endgame_plan_df,
                                            use_container_width=True,
                                            hide_index=True,
                                        )

                                    with st.expander(
                                        "Cluster allocation details"
                                    ):
                                        st.markdown(
                                            "**Portfolio cluster summary**"
                                        )
                                        st.dataframe(
                                            team_cluster_summary,
                                            use_container_width=True,
                                            hide_index=True,
                                        )
                                        st.markdown(
                                            "**Home → selected cluster travel**"
                                        )
                                        st.dataframe(
                                            team_home_cluster_df,
                                            use_container_width=True,
                                            hide_index=True,
                                        )
                                        st.markdown(
                                            "**Cluster / workload allocation**"
                                        )
                                        st.dataframe(
                                            team_allocations_df,
                                            use_container_width=True,
                                            hide_index=True,
                                        )

                                    if str(weekly_notes or "").strip():
                                        st.markdown("#### Weekly notes result")
                                        if special_request_results_df.empty:
                                            st.warning(
                                                "Notes were provided, but no requests "
                                                "could be parsed. The baseline schedule "
                                                "was kept unchanged."
                                            )
                                        else:
                                            accepted_count = int(
                                                (
                                                    special_request_results_df[
                                                        "status"
                                                    ] == "Accepted"
                                                ).sum()
                                            )
                                            rejected_count = int(
                                                (
                                                    special_request_results_df[
                                                        "status"
                                                    ] == "Rejected"
                                                ).sum()
                                            )
                                            n1, n2 = st.columns(2)
                                            n1.metric(
                                                "Accepted notes",
                                                accepted_count,
                                            )
                                            n2.metric(
                                                "Rejected notes",
                                                rejected_count,
                                            )
                                            st.dataframe(
                                                special_request_results_df.rename(
                                                    columns={
                                                        "raw_note": "Note",
                                                        "surveyor_name": "Surveyor",
                                                        "requested_date": "Date",
                                                        "location": "Requested Area",
                                                        "status": "Status",
                                                        "target_cluster": "Target Cluster",
                                                        "anchor_to_cluster_minutes": (
                                                            "Area → Cluster (Minutes)"
                                                        ),
                                                        "reason": "Result / Reason",
                                                    }
                                                ),
                                                use_container_width=True,
                                                hide_index=True,
                                            )
                                            st.caption(
                                                "Rejected notes do not modify the "
                                                "baseline schedule. Accepted notes were "
                                                "committed only after a second valid "
                                                "route trial passed the existing hard "
                                                "constraints."
                                            )

                                    st.markdown("#### Team week")
                                    st.dataframe(
                                        team_summary_df,
                                        use_container_width=True,
                                        hide_index=True,
                                    )

                                    total_team_surveys = int(
                                        team_summary_df["Surveys"].sum()
                                    )
                                    total_team_travel = int(
                                        team_summary_df[
                                            "Travel Minutes"
                                        ].sum()
                                    )
                                    total_team_survey_minutes = int(
                                        team_summary_df[
                                            "Survey Minutes"
                                        ].sum()
                                    )

                                    summary_metrics = st.columns(3)
                                    summary_metrics[0].metric(
                                        "Team surveys scheduled",
                                        total_team_surveys,
                                    )
                                    summary_metrics[1].metric(
                                        "Team survey time",
                                        f"{total_team_survey_minutes} min",
                                    )
                                    summary_metrics[2].metric(
                                        "Team transit time",
                                        f"{total_team_travel} min",
                                    )

                                    st.markdown(
                                        "#### Scheduled-building prediction reliability"
                                    )
                                    st.caption(
                                        "LOOCV MAE = mean absolute error when each "
                                        "historical building is held out once. MAE % "
                                        "is LOOCV MAE divided by this building's raw "
                                        "prediction. The ±MAE band is a practical error "
                                        "reference, not a statistical confidence interval."
                                    )
                                    if combined_team_schedule.empty:
                                        st.caption(
                                            "No scheduled buildings are available for "
                                            "reliability analysis."
                                        )
                                    else:
                                        reliability_source = team_predictions[[
                                            c for c in [
                                                "Customer Reference",
                                                "Prediction Model Used",
                                                "Prediction Confidence",
                                                "Validation MAE (Minutes)",
                                                "Validation RMSE (Minutes)",
                                                "Prediction Training Rows",
                                                "Prediction Feature Count",
                                                "Predicted Survey Duration (Minutes)",
                                            ]
                                            if c in team_predictions.columns
                                        ]].drop_duplicates(
                                            subset=["Customer Reference"]
                                        )
                                        reliability_df = combined_team_schedule[
                                            combined_team_schedule[
                                                "Sequence"
                                            ].astype(str) != "RETURN"
                                        ][[
                                            "Surveyor",
                                            "Date",
                                            "Customer Reference",
                                            "Building Name",
                                            "Postcode",
                                        ]].merge(
                                            reliability_source,
                                            on="Customer Reference",
                                            how="left",
                                        )
                                        mae = pd.to_numeric(
                                            reliability_df[
                                                "Validation MAE (Minutes)"
                                            ],
                                            errors="coerce",
                                        )
                                        pred = pd.to_numeric(
                                            reliability_df[
                                                "Predicted Survey Duration (Minutes)"
                                            ],
                                            errors="coerce",
                                        )
                                        reliability_df[
                                            "MAE as % of Prediction"
                                        ] = (100 * mae / pred).round(1)
                                        reliability_df[
                                            "Prediction - MAE (Minutes)"
                                        ] = (pred - mae).clip(
                                            lower=float(min_duration)
                                        ).round(1)
                                        reliability_df[
                                            "Prediction + MAE (Minutes)"
                                        ] = (pred + mae).round(1)
                                        def confidence_maths(row):
                                            confidence = row.get(
                                                "Prediction Confidence"
                                            )
                                            feature_count = pd.to_numeric(
                                                row.get("Prediction Feature Count"),
                                                errors="coerce",
                                            )
                                            required_medium_rows = (
                                                max(15, 5 * int(feature_count))
                                                if not pd.isna(feature_count)
                                                else 15
                                            )
                                            if confidence == "High":
                                                return (
                                                    "High because LOOCV MAE ≤ 15 "
                                                    "min and training rows ≥ 25"
                                                )
                                            if confidence == "Medium":
                                                return (
                                                    "Medium because LOOCV MAE ≤ 22 "
                                                    f"min and training rows ≥ "
                                                    f"{required_medium_rows}"
                                                )
                                            return (
                                                "Low because MAE/row thresholds for "
                                                "High or Medium were not met"
                                            )

                                        reliability_df[
                                            "Confidence Maths"
                                        ] = reliability_df.apply(
                                            confidence_maths, axis=1
                                        )
                                        st.dataframe(
                                            reliability_df,
                                            use_container_width=True,
                                            hide_index=True,
                                        )

                                    for surveyor in active_surveyors:
                                        result = team_results.get(
                                            surveyor.name
                                        )
                                        with st.expander(
                                            f"{surveyor.name} — "
                                            f"{surveyor.start_location}",
                                            expanded=True,
                                        ):
                                            shortlist = (
                                                team_shortlists.get(
                                                    surveyor.name,
                                                    pd.DataFrame(),
                                                )
                                            )
                                            st.caption(
                                                f"{len(shortlist)} sites were "
                                                "allocated to this person's Google "
                                                "candidate pool."
                                            )

                                            if result is None:
                                                st.warning(
                                                    "No candidate sites were "
                                                    "allocated."
                                                )
                                                continue

                                            st.dataframe(
                                                result.summary_dataframe(),
                                                use_container_width=True,
                                                hide_index=True,
                                            )
                                            full_person = (
                                                result
                                                .full_schedule_dataframe()
                                            )
                                            if not full_person.empty:
                                                st.dataframe(
                                                    full_person,
                                                    use_container_width=True,
                                                    hide_index=True,
                                                )

                                    if (
                                        generate_team_ai_summary
                                        and team_openai_key
                                        and not combined_team_schedule.empty
                                    ):
                                        st.markdown(
                                            "#### AI summary of the team week"
                                        )
                                        try:
                                            team_narrative = (
                                                team_planner.summarise_week(
                                                    week_summary=(
                                                        team_summary_df
                                                        .to_dict(
                                                            orient="records"
                                                        )
                                                    ),
                                                    full_schedule=(
                                                        combined_team_schedule
                                                        .to_dict(
                                                            orient="records"
                                                        )
                                                    ),
                                                    ai_site_decisions=[],
                                                    unscheduled_sites=[],
                                                    tfl_summary=(
                                                        team_tfl_summary
                                                    ),
                                                    weather_summary=(
                                                        team_weather_summary
                                                    ),
                                                    future_cluster_summary=(
                                                        team_strategy_context
                                                    ),
                                                )
                                            )
                                            st.write(team_narrative)
                                        except Exception as exc:
                                            st.warning(
                                                "Schedules were created, but "
                                                "the AI team summary failed: "
                                                f"{exc}"
                                            )

                                    team_output = io.BytesIO()
                                    with pd.ExcelWriter(
                                        team_output,
                                        engine="openpyxl",
                                    ) as writer:
                                        team_summary_df.to_excel(
                                            writer,
                                            sheet_name="Team Summary",
                                            index=False,
                                        )
                                        combined_team_schedule.to_excel(
                                            writer,
                                            sheet_name="Full Team Schedule",
                                            index=False,
                                        )
                                        write_salesforce_copy_sheet(
                                            writer,
                                            salesforce_copy_df,
                                        )
                                        team_cluster_summary.to_excel(
                                            writer,
                                            sheet_name="Cluster Summary",
                                            index=False,
                                        )
                                        pd.DataFrame(
                                            team_cluster_choices
                                        ).to_excel(
                                            writer,
                                            sheet_name="Selected Clusters",
                                            index=False,
                                        )
                                        team_endgame_plan_df.to_excel(
                                            writer,
                                            sheet_name="Endgame Plan",
                                            index=False,
                                        )
                                        team_allocations_df.to_excel(
                                            writer,
                                            sheet_name="Team Allocations",
                                            index=False,
                                        )
                                        team_home_cluster_df.to_excel(
                                            writer,
                                            sheet_name="Home Cluster Matrix",
                                            index=False,
                                        )
                                        team_google_shortlist.to_excel(
                                            writer,
                                            sheet_name="Google Shortlists",
                                            index=False,
                                        )
                                        drawing_priority_queue.to_excel(
                                            writer,
                                            sheet_name="Drawing Priority",
                                            index=False,
                                        )
                                        if not special_request_results_df.empty:
                                            special_request_results_df.to_excel(
                                                writer,
                                                sheet_name="Weekly Notes",
                                                index=False,
                                            )
                                        team_portfolio.to_excel(
                                            writer,
                                            sheet_name="Portfolio",
                                            index=False,
                                        )

                                        for surveyor in active_surveyors:
                                            result = team_results.get(
                                                surveyor.name
                                            )
                                            if result is None:
                                                continue
                                            safe_name = "".join(
                                                ch
                                                for ch in surveyor.name
                                                if ch
                                                not in r'[]:*?/\\'
                                            )[:23]
                                            sheet_name = (
                                                f"{safe_name} Week"
                                            )[:31]
                                            result.full_schedule_dataframe().to_excel(
                                                writer,
                                                sheet_name=sheet_name,
                                                index=False,
                                            )

                                    st.download_button(
                                        "Download team weekly schedule",
                                        data=team_output.getvalue(),
                                        file_name=(
                                            "Team_Survey_Week_"
                                            f"{team_week_start.isoformat()}.xlsx"
                                        ),
                                        mime=(
                                            "application/vnd.openxmlformats-"
                                            "officedocument.spreadsheetml.sheet"
                                        ),
                                        type="primary",
                                    )

                                except GoogleRoutesError as exc:
                                    st.error(str(exc))
                                except Exception as exc:
                                    st.error(
                                        f"Could not build team schedule: {exc}"
                                    )

        except Exception as exc:
            st.error(
                f"Could not process the master portfolio spreadsheet: {exc}"
            )


with tab3:
    st.subheader("Duration fallback models")
    st.dataframe(
        predictor.model_summary(),
        use_container_width=True,
        hide_index=True,
    )
    st.caption(
        "MAE is leave-one-out cross-validation error on the historical "
        "completed-survey data."
    )


